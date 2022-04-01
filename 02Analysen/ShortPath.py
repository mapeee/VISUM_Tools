# -*- coding: cp1252 -*-
#!/usr/bin/python

#-------------------------------------------------------------------------------
# Name:        Calculate shortest path from excel
# Purpose:
#
# Author:      mape
#
# Created:     23/07/2021
# Copyright:   (c) mape 2021
# Licence:     <your licence>
#-------------------------------------------------------------------------------
from openpyxl import load_workbook
import pandas as pd
from pathlib import Path
import time
start_time = time.time()
import win32com.client.dynamic

#--Path--#
path = Path.home() / 'python32' / 'python_dir.txt'
with open(path, mode='r') as f: path = f.readlines()
path = path[0] +"\\"+'VISUM_Tools'+"\\"+'ShortPath.txt'
with open(path, mode='r') as path: f = path.read().splitlines()

Network = f[0]
Excel = f[1]
PuT = 1

def get_seconds(time_str):
    hh, mm, ss = time_str.split(':')
    return int(hh) * 3600 + int(mm) * 60 + int(ss)  

def PrT_Search(V,xlsx):
    ShortPath = V.Analysis.RouteSearchPrT
    ws.cell(1,3,"tCur")
    
    for row in xlsx.iter_rows(min_row=2): #skip header
        NE = V.CreateNetElements()
        NE.Add(V.Net.Nodes.ItemByKey(row[0].value))
        NE.Add(V.Net.Nodes.ItemByKey(row[1].value))
        
        ShortPath.Execute(NE,"C",1)
        ws.cell(row[0].row,3,ShortPath.AttValue("tCur"))
        ShortPath.Clear()
        
        print("> row: "+str(row[0].row))
    
def PuT_Search(V,xlsx):
    global List
    ShortPath = V.Analysis.RouteSearchPuT
    ws.cell(1,3,"Dep")
    ws.cell(1,4,"Arr")
    ws.cell(1,5,"Time_min")
    ws.cell(1,6,"Dist_km")
    ws.cell(1,7,"Transfers")
    ws.cell(1,8,"Lines")
    
    for row in xlsx.iter_rows(min_row=2): #skip header
        ShortPath.Clear()
        NE = V.CreateNetElements()
        NE.Add(V.Net.StopAreas.ItemByKey(row[0].value))
        NE.Add(V.Net.StopAreas.ItemByKey(row[1].value))
        
        ShortPath.Execute(NE,"OV","08:00:00",1,True)
        List=V.Lists.CreatePuTPathSearchLegList
        List.AddColumn("Dep")
        List.AddColumn("Arr")
        List.AddColumn("Dist")
        List.AddColumn("VEHJOURNEY\LINEROUTE\LINE\LINIENNUMMER")
        List = pd.DataFrame(List.SaveToArray())
        dep = get_seconds(List.iloc[0][0])
        arr = get_seconds(List.iloc[-1][1])
        if dep > arr: arr+=60*60*60  ##travel over midnight
        tt = int((arr-dep)/60)
        ws.cell(row[0].row,3,List.iloc[0][0])
        ws.cell(row[0].row,4,List.iloc[-1][1])
        ws.cell(row[0].row,5,tt)
        ws.cell(row[0].row,6,sum(List[2]))
        ws.cell(row[0].row,7,len(List[List[3]!=""])-1)
        ws.cell(row[0].row,8,List[List[3]!=""][3].str.cat(sep=', '))
        
        print("> row: "+str(row[0].row))

def VISUM_open(Net):
    VISUM = win32com.client.dynamic.Dispatch("Visum.Visum.22")
    VISUM.loadversion(Net)
    VISUM.Filters.InitAll()
    return VISUM 


#calculate
wb = load_workbook(filename = Excel)
ws = wb.worksheets[0]

VISUM = VISUM_open(Network)

if PuT ==1: PuT_Search(VISUM,ws)
else: PrT_Search(VISUM,ws)


#END
VISUM = False
wb.save(Excel)
wb.close()
print("> finished after "+str(int(time.time()-start_time))+" seconds")