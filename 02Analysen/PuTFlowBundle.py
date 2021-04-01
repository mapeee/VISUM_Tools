# -*- coding: cp1252 -*-
#!/usr/bin/python

#-------------------------------------------------------------------------------
# Name:        calculating FlowBundles to value PuT-Scenarios
# Purpose:
#
# Author:      mape
#
# Created:     05/01/2021
# Copyright:   (c) mape 2021
# Licence:     <your licence>
#-------------------------------------------------------------------------------
import win32com.client.dynamic
import numpy as np
import openpyxl
import time
start_time = time.time()

from pathlib import Path
path = Path.home() / 'python32' / 'python_dir.txt'
f = open(path, mode='r')
for i in f: path = i
path = Path.joinpath(Path(path),'VISUM_Tools','FlowBundle.txt')
f = path.read_text()
f = f.split('\n')

#--Path--#
Network = f[0]
XLSX = f[1]

#--Excel--#
XLSX = openpyxl.load_workbook(XLSX)
HstBer = XLSX.active
#Excel results
Resluts_xlsx = XLSX.create_sheet(index = 2 , title = "results")
Resluts_xlsx.cell(1,1,"From_NO")
Resluts_xlsx.cell(1,2,"From_Name")
Resluts_xlsx.cell(1,3,"TO_NO")
Resluts_xlsx.cell(1,4,"TO_Name")
Resluts_xlsx.cell(1,5,"volume")
Resluts_xlsx.cell(1,6,"transfers")


def VISUM_open(Net):
    VISUM = win32com.client.dynamic.Dispatch("Visum.Visum.20")
    VISUM.loadversion(Net)
    return VISUM

# Network =r'I:\Verkehrsmodell\10TestArea\01Versionsdateien\Test_OEV_Umstieg.ver'
VISUM = VISUM_open(Network)

MyFlowBundle = VISUM.Net.DemandSegments.ItemByKey("PV_OV_mFV_Nachm").FlowBundle
MyFlowBundle.Clear()

MyFlowBundle.DemandSegments = ("PV_OV_mFV_Vorm,PWV_OV_Vorm,"
"PV_OV_mFV_Nachm,PWV_OV_Nachm,"
"PV_OV_mFV_Nacht,PWV_OV_Nacht,"
"PV_OV_mFV_Rest,PWV_OV_Rest")

i = 2 ##headline in first row
for origin in range(1,HstBer.max_row+1):
    if origin < 2:continue ##< 2 due to headline

    for destination in range(1,HstBer.max_row+1):
        if destination < 2:continue ##< 2 due to headline
        if origin == destination:continue
        
        if origin <12 and destination <12: continue ##nur Landwehr und Altona
        
        print(HstBer.cell(origin,2).value)
        print(HstBer.cell(destination,2).value) 
        
        NetElementContainer = VISUM.CreateNetElements()
        NetElementContainer.Add(VISUM.Net.StopAreas.ItemByKey(HstBer.cell(origin,1).value))
        NetElementContainer.Add(VISUM.Net.StopAreas.ItemByKey(HstBer.cell(destination,1).value))

        MyFlowBundle.Execute(NetElementContainer)
        
        
        Volume = round(sum(np.array(VISUM.Net.StopPoints.GetMultiAttValues("PassOriginFlowBundle(AP)"))[:,1]))
        Boarding = round(sum(np.array(VISUM.Net.StopPoints.GetMultiAttValues("PassBoardFlowBundle(AP)"))[:,1]))
        Transfer = Boarding - Volume
        
        Resluts_xlsx.cell(i,1,HstBer.cell(origin,1).value)
        Resluts_xlsx.cell(i,2,HstBer.cell(origin,2).value)
        Resluts_xlsx.cell(i,3,HstBer.cell(destination,1).value)
        Resluts_xlsx.cell(i,4,HstBer.cell(destination,2).value)
        Resluts_xlsx.cell(i,5,Volume)
        Resluts_xlsx.cell(i,6,Transfer)
        
        MyFlowBundle.Clear()
        
        i+=1
        

#End
XLSX.save(r'C:\Users\peter\Desktop\test.xlsx')
XLSX.close()

Sekunden = int(time.time() - start_time)
print("--finished after ",Sekunden,"seconds--")