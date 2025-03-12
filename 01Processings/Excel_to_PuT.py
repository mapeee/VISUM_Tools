# -*- coding: cp1252 -*-
#!/usr/bin/python

#-------------------------------------------------------------------------------
# Name:        PuT from Excel to PTV Visum
# Author:      mape
# Created:     08/04/2021 // 15/05/2023
# Copyright:   (c) mape 2021
# Licence:     <your licence>
#-------------------------------------------------------------------------------

import win32com.client.dynamic
import openpyxl
from pathlib import Path
f = open(Path.home() / 'python32' / 'python_dir.txt', mode='r')
path = list(f)[0]
path = Path.joinpath(Path(path), 'VISUM_Tools', 'Excel_to_PuT.txt')
f = path.read_text().split('\n')

Network = f[0]
XLSX = f[1]

#--Parameter--#
LineName = "A2"
TSys = "Bus"
Direction = "A5"
StopsFrom = "F18"


def CreateHeadway(Headway, Dep, Dep_next, No, NoVJ):
    HeadwayJourneys = int(((Dep_next-Dep)/Headway)-1)
    for HeadwayJourney in range(HeadwayJourneys):
        NoHead = int(str(No)+str(HeadwayJourney+1))
        DepHead = Dep + (Headway*(HeadwayJourney+1))
        CreateVehJourney(Visum, TimeProfile, NoHead, NoVJ, DepHead)

def CreateLine(_Visum, _XLSX, __LineName, _TSys):
    _LineName = _XLSX[__LineName].value
    try:
        Line = _Visum.Net.Lines.ItemByKey(_LineName)
        print("> Line", _LineName, "already exists")
        return Line
    except:
        print("> creating Line ", _LineName)
        Line = _Visum.Net.AddLine(_LineName, _TSys)
        return Line

def CreateLineRoute(_Visum, _XLSX, _Line, _Stops, _Direction):
    if _Line.LineRoutes.Count == 0:
        _No = 1
    else:
        _No = int(_Line.AttValue(r"MAX:LINEROUTES\ID")) + 1
    _RouteName = str(_No)
    _Direction = _XLSX[_Direction].value
    if _Direction == "Richtung 1": _Direction = "H"
    if _Direction == "Richtung 2": _Direction = "R"
    if _Direction not in ["H", "R"]:
        print("> wrong direction:", _Direction)
        raise
    RouteSearch = Visum_para(_Visum)
    LineRoute = _Visum.Net.AddLineRoute(_RouteName, _Line, _Direction, _Stops, RouteSearch)
    if LineRoute == None:
        print("> no LineRoute added")
        raise
    return LineRoute

def CreateTimeProfile(_Visum, _Route):
    TProfile = _Visum.Net.AddTimeProfile("1", _Route)
    # for Item in TProfile.TimeProfileItems:
    #     Index = Item.AttValue("Index")
    #     prevarr = Item.AttValue("PREVTIMEPROFILEITEM\ARR")
    #     if Index==1:continue
    #     arr = prevarr+Times[int(Index-2)]
    #     Item.SetAttValue("arr",arr)
    return TProfile    

def CreateVehJourney(_Visum, TimeProfile, No, NoVJ, Dep):
    VJ = _Visum.Net.AddVehicleJourney(NoVJ,TimeProfile)
    VJ.SetAttValue("Dep",Dep)
    VJ.SetAttValue("Name",str(No))
    Operator = VJ.AttValue('LINEROUTE\LINE\OPERATORNO')
    VJ.SetAttValue('OPERATORNO',Operator)
    return VJ

def DepartureTime(_XLSX, Col):
    for row in range(10,_XLSX.max_row+10):
        Time = _XLSX.cell(row,Col).value
        if Time in [None,"|","S"]:continue
        Departure = Times_to_seconds(Time,None)
        break
    return Departure

def Stop_Container(_Visum, _XLSX, _StopsFrom):
    Stops = _Visum.CreateNetElements()
    _row = _XLSX[_StopsFrom].row
    _col = _XLSX[_StopsFrom].column
    for row in range(_row,_XLSX.max_row+10):
        Stop_No = _XLSX.cell(row, _col).value
        if Stop_No == None:
            return Stops
        if Stop_No == _XLSX.cell(row-1, _col).value:
            print("> Identical StopID one after the other:", Stop_No)
            raise
        try:
            _Stop = _Visum.Net.StopPoints.ItemByKey(Stop_No)
        except:
            print("> StopID", Stop_No, "not found")
            raise
        Stops.Add(_Stop)

def Times_to_seconds(Time_from, Time_to):
    if Time_from == None:
        return None
    if Time_to == None:
        departure = (int(str(Time_from)[-2:])+(int(str(Time_from)[:-2])*60))*60
        return departure
    minutes_from = int(str(Time_from)[-2:])+(int(str(Time_from)[:-2])*60)
    minutes_to = int(str(Time_to)[-2:])+(int(str(Time_to)[:-2])*60)
    seconds = (minutes_to - minutes_from)*60
    return seconds

def Visum_open(Net):
    _Visum = win32com.client.dynamic.Dispatch("Visum.Visum.24")
    _Visum.loadversion(Net)
    _Visum.Filters.InitAll()
    return _Visum

def Visum_para(_Visum):
    Routing = _Visum.IO.CreateNetReadRouteSearchTsys()
    Routing.SetAttValue("ChangeLinkTypeOfOpenedLinks", False)
    Routing.SetAttValue("IncludeBlockedTurns", True)
    Routing.SetAttValue("HowToHandleIncompleteRoute", 2) # search shortest path
    Routing.SetAttValue("LinkTypeForInsertedLinksReplacingMissingLinks", 1)
    Routing.SetAttValue("ShortestPathCriterion", 1) # link length (1 = travel time; 3 = link length)
    Routing.SetAttValue("MaxDeviationFactor", 50)
    Routing.SetAttValue("WhatToDoIfShortestPathNotFound", 1) # open link / turn
    Routing.SetAttValue("LinkTypeForOpenedLinks", 1)
    Routing.SetAttValue("LinkTypeForInsertedLinksReplacingMissingShortestPaths", 1)
    Routing.SetAttValue("WhatToDoIfStopPointIsBlocked", 2) # open stopPoint
    Routing.SetAttValue("WhatToDoIfStopPointNotFound", 0) # do not read lineroute
    
    return Routing

#--Processing--#
Visum = Visum_open(Network)
wb = openpyxl.load_workbook(XLSX, data_only = True)

for Excel in wb.worksheets:
    if Excel.title == "Tabelle1":
        continue
    Line = CreateLine(Visum, Excel, LineName, TSys)
    try:
        Stops = Stop_Container(Visum, Excel, StopsFrom)
        LineRoute = CreateLineRoute(Visum, Excel, Line, Stops, Direction)
        TimeProfile = CreateTimeProfile(Visum, LineRoute)
    except:
        print("> error")
    print()

# VHNo = int(Visum.Net.VehicleJourneys.GetMultipleAttributes(["NO"])[-1][0]) + 1
# Direct = "R"

# Elements = Excel

# No=Dep=Stops=Times=LineRoute=TimeProfile=VehJourney = None

# for Journey in range(7,Elements.max_column+1):
#     # No = Elements.cell(9,Journey).value
#     # print("> Journey: "+str(No))
#     # if HeadwayCheck(Elements.cell(10,Journey).value) != None:
#     #     print("> Headway Journey: "+str(No))
#     #     Headway = HeadwayCheck(Elements.cell(10,Journey).value)
#     #     Dep_next = DepartureTime(Elements,Journey+1)
#     #     CreateHeadway(Headway, Dep, Dep_next, No, VHNo)
#     #     VHNo+=1
#     #     continue          
#     Dep = DepartureTime(Elements,Journey)
#     Stops, Times = Stop_Container(Visum, Elements, Journey)
#     LineRoute = CreateLineRoute(Visum, Line, No, Stops, Direct)
#     TimeProfile = CreateTimeProfile(Visum, LineRoute, Times)
#     VehJourney = CreateVehJourney(Visum, TimeProfile, No, VHNo, Dep)
#     VHNo+=1

#End
print("> finished")