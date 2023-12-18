# -*- coding: cp1252 -*-
#!/usr/bin/python

#-------------------------------------------------------------------------------
# Name:        importing excel to matrice
# Purpose:
# Author:      mape
# Created:     31/01/2023
# Copyright:   (c) mape 2023
# Licence:     <your licence>
#-------------------------------------------------------------------------------

'''
Help

Import of Excel-OD-Date to VISUM Matrice
From/To Zones are referenced by index number instead of zone number!
'''

import win32com.client.dynamic
import openpyxl
VISUM = win32com.client.dynamic.Dispatch("Visum.Visum.22")

#--Parameters--#
Net = r"C:\Users\peter\hvv.de\S-GR-Modellierung - 10 Verkehrsmodell\1060 Projekte\Sonstige\MobiHam\MobiHam_2022.ver"
Excel = r"C:\Users\peter\OneDrive - hvv.de\Desktop\MobiHam.xlsx"

mat_nr = 3 #Number of VISUM-matrice
orig_col = 5 #column number of origins (A1 = 1)
desti_col = 6
header = 1 #0 = no; 1 = yes

#--Process--#
VISUM.loadversion(Net)
mat = VISUM.Net.Matrices.ItemByKey(mat_nr)
XLSX = openpyxl.load_workbook(Excel).active

for row in range(1+header,XLSX.max_row+header):
    orig = XLSX.cell(row,orig_col).value
    desti = XLSX.cell(row,desti_col).value
    if orig == 0 or desti == 0:continue
    mat.SetValue(orig,desti,1)
    
print("> finish")



