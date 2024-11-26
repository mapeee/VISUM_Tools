# -*- coding: cp1252 -*-
#!/usr/bin/python

#-------------------------------------------------------------------------------
# Name:        PuT from Excel to PTV Visum
# Author:      mape
# Created:     08/04/2021 // 15/05/2023
# Copyright:   (c) mape 2021
# Licence:     <your licence>
#-------------------------------------------------------------------------------

import win32com.client.dynamic
import sys
import openpyxl
import pandas as pd
from pathlib import Path
f = open(Path.home() / 'python32' / 'python_dir.txt', mode='r')
path = list(f)[0]
path = Path.joinpath(Path(path), 'VISUM_Tools', 'Excel_to_SystemRoute.txt')
f = path.read_text().split('\n')
sys.path.insert(0,f[2])
from VisumPy.helpers import SetMulti

#--Parameter--#
TSys = "D2"
Name = "C2"
StopsFrom = "A2"
ValuesFrom = "B2"
_UDA = "AFZ_OEV_2024"

def CreateSystemRoute(_Visum, _XLSX, _StopsFrom, _Name, _TSys, _RouteSearch):
    __Name = _XLSX[_Name].value
    __TSys = _XLSX[_TSys].value
    _Stops = Stop_Container(_Visum, _XLSX, _StopsFrom)
    try:
        SystemRoute = _Visum.Net.SystemRoutes.ItemByKey(__Name)
        print("> SystemRoute", __Name, "already exists")
        return SystemRoute
    except:
        print("> SystemRoute ", __Name, "created")
        SystemRoute = _Visum.Net.AddSystemRoute(__Name, __TSys, _Stops, _RouteSearch)
        return SystemRoute

def SRValues(_SystemRoute, _XLSX, _ValuesFrom, _StopsFrom):
    Stop_Col = _XLSX.cell(_XLSX[_StopsFrom].row-1, _XLSX[_StopsFrom].column).value
    Value_Col = _XLSX.cell(_XLSX[_ValuesFrom].row-1, _XLSX[_ValuesFrom].column).value
    df = pd.DataFrame(_XLSX.values, columns=next(_XLSX.values))[[Stop_Col, Value_Col]]

    Value = 0
    l = []
    for SRElement in _SystemRoute.SystemRouteItems.GetAll:
        if SRElement.AttValue(r"OUTLINK\NO") == None:
            l.append(0)
            break
        if SRElement.AttValue("STOPPOINTNO") == None:
            l.append(Value)
        else:
            HPNO = int(SRElement.AttValue("STOPPOINTNO"))
            if len(df[df[Stop_Col]==HPNO]) == 0:
                l.append(Value)
                continue
            if df[df[Stop_Col]==HPNO][Value_Col].iloc[0] == 0:
                l.append(Value)
                continue
            Value = df[df[Stop_Col]==HPNO][Value_Col].iloc[0]
            l.append(Value)
    
    Items = _SystemRoute.SystemRouteItems
    SetMulti(Items, _UDA, l, False)
                
def Stop_Container(_Visum, _XLSX, _StopsFrom):
    Stops = _Visum.CreateNetElements()
    _row = _XLSX[_StopsFrom].row
    _col = _XLSX[_StopsFrom].column
    for row in range(_row,_XLSX.max_row+10):
        Stop_No = _XLSX.cell(row, _col).value
        if Stop_No == None:
            return Stops
        if Stop_No == _XLSX.cell(row-1, _col).value:
            print("> Identical StopID one after the other:", Stop_No)
            raise
        try:
            _Stop = _Visum.Net.StopPoints.ItemByKey(Stop_No)
        except:
            print("> StopID", Stop_No, "not found")
            raise
        Stops.Add(_Stop)

def Visum_open(Net):
    _Visum = win32com.client.dynamic.Dispatch("Visum.Visum.24")
    _Visum.loadversion(Net)
    _Visum.Filters.InitAll()
    return _Visum

def Visum_para(_Visum):
    Routing = _Visum.IO.CreateNetReadRouteSearchTsys()
    Routing.SetAttValue("HowToHandleIncompleteRoute", 2) # search shortest path
    Routing.SetAttValue("IncludeBlockedLinks", False)
    Routing.SetAttValue("IncludeBlockedTurns", False)
    Routing.SetAttValue("ShortestPathCriterion", 1) # link length (1 = travel time; 3 = link length)
    return Routing

#--Processing--#
Visum = Visum_open(f[0])
RouteSearch = Visum_para(Visum)
wb = openpyxl.load_workbook(f[1], data_only = True)
for ws in wb.worksheets:
    SystemRoute = CreateSystemRoute(Visum, ws, StopsFrom, Name, TSys, RouteSearch)
    SRValues(SystemRoute, ws, ValuesFrom, StopsFrom)

#End
print("> finished")