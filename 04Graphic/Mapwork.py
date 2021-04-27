# -*- coding: cp1252 -*-
#!/usr/bin/python

#-------------------------------------------------------------------------------
# Name:        PuT-Maps in VISUM
# Purpose:
#
# Author:      mape
#
# Created:     08/04/2021
# Copyright:   (c) mape 2021
# Licence:     <your licence>
#-------------------------------------------------------------------------------

import win32com.client.dynamic
import openpyxl
from random import randrange
import time
start_time = time.time()

from pathlib import Path
path = Path.home() / 'python32' / 'python_dir.txt'
f = open(path, mode='r')
for i in f: path = i
path = Path.joinpath(Path(path),'VISUM_Tools','Mapwork.txt')
f = path.read_text()
f = f.split('\n')

def Excel(path):
    XLSX = path
    XLSX = openpyxl.load_workbook(XLSX)
    return XLSX

def VISUM_para(VISUM):
    Routing = VISUM.CreateNetReadRouteSearchTsys()
    Routing.SetAttValue("ChangeLinkTypeOfOpenedLinks",False)
    Routing.SetAttValue("IncludeBlockedTurns",True)
    Routing.SetAttValue("HowToHandleIncompleteRoute", 2) ##search shortest path
    Routing.SetAttValue("LinkTypeForInsertedLinksReplacingMissingShortestPaths",1)
    Routing.SetAttValue("ShortestPathCriterion",3) ##link length (1 = travel time; 3 = link length)
    Routing.SetAttValue("MaxDeviationFactor",3)
    Routing.SetAttValue("WhatToDoIfShortestPathNotFound",1) ##open turns
    Routing.SetAttValue("LinkTypeForInsertedLinksReplacingMissingShortestPaths",1)
    Routing.SetAttValue("WhatToDoIfStopPointIsBlocked", 2) ##Open the StopPoint
    Routing.SetAttValue("WhatToDoIfStopPointNotFound", 0) ##do not read lineroute
    return Routing

def AddNodeStop(VISUM,Excel,row):
    Nr = Excel.cell(row,1).value
    Name = Excel.cell(row,2).value
    X = Excel.cell(row,3).value
    Y = Excel.cell(row,4).value
    Helper = Excel.cell(row,5).value
    
    print(Name)

    #Node Stops
    Node = VISUM.Net.AddNode(Nr,X,Y)
    Node.SetAttValue("Name",Name)
    Stop = VISUM.Net.AddStop(Nr,X,Y)
    Stop.SetAttValue("Name",Name)
    StopArea = VISUM.Net.AddStopArea(Nr,Stop,Node)
    StopArea.SetAttValue("Name",Name)
    StopPoint = VISUM.Net.AddStopPointOnNode(Nr,StopArea,Node)
    StopPoint.SetAttValue("Name",Name)
    
    VISUM.Graphic.Autozoom(Node)
 
    #Helper
    if Helper == 0: return
    Node = VISUM.Net.AddNode(int(str(Nr)+"001"),X+3,Y) #east
    Node = VISUM.Net.AddNode(int(str(Nr)+"002"),X+3,Y-3) #southeast
    Node = VISUM.Net.AddNode(int(str(Nr)+"003"),X,Y-3) #south
    Node = VISUM.Net.AddNode(int(str(Nr)+"004"),X-3,Y-3) #southwest
    Node = VISUM.Net.AddNode(int(str(Nr)+"005"),X-3,Y) #west
    Node = VISUM.Net.AddNode(int(str(Nr)+"006"),X-3,Y+3) #northwest
    Node = VISUM.Net.AddNode(int(str(Nr)+"007"),X,Y+3) #north
    Node = VISUM.Net.AddNode(int(str(Nr)+"008"),X+3,Y+3) #northeast
    
    
def AddHelperLink(VISUM,Excel,row):
    Nr = Excel.cell(row,1).value
    Helper = Excel.cell(row,5).value
    if Helper == 0:return
    
    #Links
    VISUM.Net.AddLink(int(str(Nr)+"001"),int(str(Nr)+"001"),int(str(Nr)+"002"),1)
    VISUM.Net.AddLink(int(str(Nr)+"002"),int(str(Nr)+"002"),int(str(Nr)+"003"),1)
    VISUM.Net.AddLink(int(str(Nr)+"003"),int(str(Nr)+"003"),int(str(Nr)+"004"),1)
    VISUM.Net.AddLink(int(str(Nr)+"004"),int(str(Nr)+"004"),int(str(Nr)+"005"),1)
    VISUM.Net.AddLink(int(str(Nr)+"005"),int(str(Nr)+"005"),int(str(Nr)+"006"),1)
    VISUM.Net.AddLink(int(str(Nr)+"006"),int(str(Nr)+"006"),int(str(Nr)+"007"),1)
    VISUM.Net.AddLink(int(str(Nr)+"007"),int(str(Nr)+"007"),int(str(Nr)+"008"),1)
    VISUM.Net.AddLink(int(str(Nr)+"008"),int(str(Nr)+"008"),int(str(Nr)+"001"),1)
    
    VISUM.Net.AddLink(int(str(Nr)+"010"),int(str(Nr)+"005"),Nr,1)
    VISUM.Net.AddLink(int(str(Nr)+"011"),int(str(Nr)+"001"),Nr,1)

    
def AddConnections(VISUM,Excel):
    for row in range(2,Stops.max_row+1):
        from_help = Excel.cell(row,5).value
        to_help = Excel.cell(row+1,5).value
        from_no = Excel.cell(row,1).value
        to_no = Excel.cell(row+1,1).value
        link_no = int(str(from_no)+"012")
        
        if to_no == None: return
        
        if from_help == 1: from_no = int(str(from_no)+"001")
        if to_help == 1: to_no = int(str(to_no)+"005")
         
        VISUM.Net.AddLink(link_no,from_no,to_no,1)
    
def AddLine(VISUM,Excel,col):
    Name = col[0].value
    Line = VISUM.Net.AddLine(Name,"B")
    print(Name)
    
    StopContainer = VISUM.CreateNetElements()
    for row in range(2,Stops.max_row+1):
        Stop = Excel.cell(row,1).value
        if col[row-1].value == 1:StopContainer.Add(VISUM.Net.StopPoints.ItemByKey(Stop))
        if col[row-1].value == 0:StopContainer.Add(VISUM.Net.Nodes.ItemByKey(int(str(Stop)+"003")))
    dirTo = VISUM.Net.Directions.GetAll[0]
    RouteSearch = VISUM_para(VISUM)
    Route = VISUM.Net.AddLineRoute("Route1",Line,dirTo,StopContainer,RouteSearch)
    
    TProfile = VISUM.Net.AddTimeProfile("Prof1",Route)
    i = 1
    for Arr in TProfile.TimeProfileItems:
        try:
            Arr.SetAttValue("Arr",300*i)
            i+=1
        except:pass
        
    VISUM.Net.AddVehicleJourney(randrange(1000),TProfile)
    
    VISUM.Graphic.Autozoom(Route)


#--Parameter--#
XLSX = Excel(f[0])
Stops = XLSX.active
VISUM = win32com.client.dynamic.Dispatch("Visum.Visum.20")

#--Build Network--#
for row in range(2,Stops.max_row+1):
    AddNodeStop(VISUM, Stops, row)
    AddHelperLink(VISUM,Stops, row)

AddConnections(VISUM,Stops)

#--Add Lines--#
for col in Stops.iter_cols(6, Stops.max_column):
    AddLine(VISUM,Stops,col)

#--output--#
VISUM.Net.GraphicParameters.Open(f[1],True)
VISUM.Filters.LinkFilter().AddCondition("OP_None",False,"Length", "GreaterVal", 0.0300)

#End
XLSX.close()
seconds = int(time.time() - start_time)
print("--finished after ",seconds,"seconds--")