# -*- coding: cp1252 -*-
#!/usr/bin/python

#-------------------------------------------------------------------------------
# Name:        Demand-Data from Excel/FAN to PTV VISUM
# Author:      mape
# Created:     11/12/2023
# Copyright:   (c) mape 2023
# Licence:     CC-BY-4.0
#-------------------------------------------------------------------------------

import win32com.client.dynamic
import openpyxl

from pathlib import Path
path = Path.home() / 'python32' / 'python_dir.txt'
path = open(path, mode='r').readlines()
path = path[0] +"\\"+'VISUM_Tools'+"\\"+'FAN_VISUM.txt'
f = open(path, mode='r').read().splitlines()
Network = f[0]
XLSX = f[1]

#--Parameter--#
Sheet = "Kanten"
From = 1
To = 3
Demand = 8
Lines = 9
Routes = 10
Direction = 11

def FAN_VISUM(f,t,v,r):
    Items = r.LineRouteItems.GetAll
    for Item in Items:
        FromStop = Item.AttValue("STOPPOINTNO")
        ToStop = Item.AttValue(r"NEXTROUTEPOINT\STOPPOINTNO")
        if FromStop is None:
            FromStop = Item.AttValue(r"PREVROUTEPOINT\STOPPOINTNO")
        if ToStop is None:
            Item.SetAttValue("ADDVAL",0)
            break
        if FromStop == f and ToStop == t:
            Item.SetAttValue("ADDVAL",v)
            if Item.AttValue(r"NEXTLINEROUTEITEM\STOPPOINTNO") is not None:
                return

#--Process--#
VISUM = win32com.client.dynamic.Dispatch("Visum.Visum.23")
VISUM.loadversion(Network)

Data = openpyxl.load_workbook(XLSX)[Sheet]
for row in range(2,Data.max_row+1):
    FromStop = Data.cell(row,From).value
    ToStop = Data.cell(row,To).value
    Vol = Data.cell(row,Demand).value
    LinesNo = Data.cell(row,Lines).value
    RoutesNo = Data.cell(row,Routes).value
    Direct = Data.cell(row,Direction).value
    if RoutesNo == 0:
        print("error")
        continue
    
    for Line in LinesNo.split(","):
        if len(LinesNo.split(","))==1:
            Route = VISUM.Net.LineRoutes.ItemByKey(Line, Direct, RoutesNo)
            FAN_VISUM(FromStop,ToStop,Vol,Route)
            break
        LineRoutes = VISUM.Net.Lines.ItemByKey(Line.strip()).LineRoutes.GetAll
        for LineRoute in LineRoutes:
            FAN_VISUM(FromStop,ToStop,Vol,LineRoute)



