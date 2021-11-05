# -*- coding: cp1252 -*-
#!/usr/bin/python

#-------------------------------------------------------------------------------
# Name:        Networkelements from Excel to PTV VISUM
# Purpose:
#
# Author:      mape
#
# Created:     08/04/2021
# Copyright:   (c) mape 2021
# Licence:     <your licence>
#-------------------------------------------------------------------------------

import win32com.client.dynamic
import openpyxl
import time
start_time = time.time()
from pathlib import Path
f = open(Path.home() / 'python32' / 'python_dir.txt', mode='r')
for i in f: path = i
path = Path.joinpath(Path(path),'VISUM_Tools','Excel_to_VISUM.txt')
f = path.read_text().split('\n')

global addition_time, extension_factor

def Excel(path):
    XLSX = path
    XLSX = openpyxl.load_workbook(XLSX)
    return XLSX

def LineLine(VISUM, LineName, VSys):
    try: Line = VISUM.Net.Lines.ItemByKey(LineName)
    except: Line = VISUM.Net.AddLine(LineName,VSys) ##Add linename in TSys
    return Line

def Lines(VISUM,Excel):
    for row in range(2,Excel.max_row+1): #no header, from A2
        LineName = Excel.cell(row,1).value #column A
        RouteName = Excel.cell(row,2).value #column B
        VSys = Excel.cell(row,3).value #column C
        Orig = Excel.cell(row,4).value
        Desti = Excel.cell(row,5).value
        Direct = Excel.cell(row,6).value
        Headway = Excel.cell(row,7).value
        Index = Excel.cell(row,8).value
        Desti_PrT = Excel.cell(row,9).value
        Orig_PrT = Excel.cell(row,10).value
        
        print("-> Importing Line "+LineName+ " "+Direct)
        
        Line = LineLine(VISUM, LineName, VSys)
        Route = LineRoute(VISUM, Line, RouteName, Direct, Orig, Desti)
        Profile = TimeProfile(VISUM, Route, Orig_PrT, Desti_PrT)
        VehJourneys(VISUM, Profile, Headway, Index)
        print (" ")
    
def LineRoute(VISUM, Line, RouteName, Direct, Orig, Desti):
    RouteName = RouteName+"_"+Direct
    STOPS = Stop_Container(VISUM, Orig, Desti)
    RouteSearch = VISUM_para(VISUM)
    Route = VISUM.Net.AddLineRoute(RouteName,Line,Direct,STOPS,RouteSearch) ##
    return Route

def Node_Container(VISUM, Orig, Desti):
    Container = VISUM.CreateNetElements() ##Container including stoppoints of lineroute
    Container.Add(VISUM.Net.Nodes.ItemByKey(Orig))
    Container.Add(VISUM.Net.Nodes.ItemByKey(Desti))
    return Container

def PrTShortPath(VISUM, Orig_PrT, Desti_PrT):
    ShortPath = VISUM.Analysis.RouteSearchPrT
    NODES = Node_Container(VISUM, Orig_PrT, Desti_PrT)
    ShortPath.Execute(NODES,"C",1) ##1 = current speed
    tcur = (int(ShortPath.AttValue("tCur"))*extension_factor) + addition_time
    ShortPath.Clear()
    return tcur

def Stop_Container(VISUM, Orig, Desti):
    Container = VISUM.CreateNetElements() ##Container including stoppoints of lineroute
    Container.Add(VISUM.Net.StopPoints.ItemByKey(Orig))
    Container.Add(VISUM.Net.StopPoints.ItemByKey(Desti))
    return Container

def Stops(VISUM,Excel):
    for row in range(2,Excel.max_row+1): #no header, from A2
        Nr = Excel.cell(row,1).value #column A
        Name = Excel.cell(row,2).value #column B
        Node = VISUM.Net.Nodes.ItemByKey(Nr)
        X = Node.AttValue("XCoord")
        Y = Node.AttValue("YCoord")
        #Add Stops
        try:
            Node.SetAttValue("Name",Name)
            Stop = VISUM.Net.AddStop(Nr,X,Y)
            Stop.SetAttValue("Name",Name)
            StopArea = VISUM.Net.AddStopArea(Nr,Stop,Node)
            StopArea.SetAttValue("XCoord",X)
            StopArea.SetAttValue("YCoord",Y)
            StopArea.SetAttValue("Name",Name)
            StopPoint = VISUM.Net.AddStopPointOnNode(Nr,StopArea,Node)
            StopPoint.SetAttValue("Name",Name)
            StopPoint.SetAttValue("AddVal2",1)
            print("added: "+Name)
        except: print("error at: "+(str(Nr)))
        
def TimeProfile(VISUM, Route, Orig_PrT, Desti_PrT):
    TProfile = VISUM.Net.AddTimeProfile("1",Route)
    tPT = PrTShortPath(VISUM, Orig_PrT, Desti_PrT)
    if tPT == addition_time:
        print("-> no travel time")
        tPT+=600
        for i in TProfile.TimeProfileItems:
            if i.AttValue("arr") >0: i.SetAttValue("arr",tPT)
    else:
        for i in TProfile.TimeProfileItems:
            if i.AttValue("arr") >0: i.SetAttValue("arr",tPT)
    return TProfile
  
def VehJourneys(VISUM, TimeProfile, Headway, Index):
    # Journey = VISUM.Net.AddVehicleJourney(float(Index),TimeProfile)
    times = list(enumerate(range(0,86400,Headway*60))) ##86400 == one day 24h
    for slot in times:
        try:
            Journey = VISUM.Net.AddVehicleJourney(slot[0]+Index,TimeProfile)
            Journey.SetAttValue("Dep",slot[1])
        except:
            print("-> Missing journeys")
            break
      
def VISUM_open(Net):
    VISUM = win32com.client.dynamic.Dispatch("Visum.Visum.20")
    VISUM.loadversion(Net)
    VISUM.Filters.InitAll()
    return VISUM

def VISUM_para(VISUM):
    Routing = VISUM.CreateNetReadRouteSearchTsys()
    Routing.SetAttValue("ChangeLinkTypeOfOpenedLinks",False)
    Routing.SetAttValue("IncludeBlockedTurns",False)
    Routing.SetAttValue("HowToHandleIncompleteRoute", 2) ##search shortest path
    Routing.SetAttValue("LinkTypeForInsertedLinksReplacingMissingShortestPaths",5)
    Routing.SetAttValue("ShortestPathCriterion",1) ##link length (1 = travel time; 3 = link length)
    Routing.SetAttValue("MaxDeviationFactor",8)
    Routing.SetAttValue("WhatToDoIfShortestPathNotFound",2) ##insert link
    Routing.SetAttValue("LinkTypeForInsertedLinksReplacingMissingShortestPaths",5)
    Routing.SetAttValue("WhatToDoIfStopPointIsBlocked", 0) ##do not read lineroute
    Routing.SetAttValue("WhatToDoIfStopPointNotFound", 0) ##do not read lineroute
    return Routing

#--Parameter--#
Network = f[0]
VISUM = VISUM_open(Network)

XLSX = Excel(f[1])
Elements = XLSX.active

#--Stops--#
# Stops(VISUM,Elements)

#--Lines--#
addition_time = 300 ##5 minutes * 60 seconds
extension_factor = 1.4
Lines(VISUM,Elements)

#End
XLSX.close()
seconds = int(time.time() - start_time)
print("-> finished after ",seconds,"seconds")