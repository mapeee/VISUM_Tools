# -*- coding: cp1252 -*-
#!/usr/bin/python

#-------------------------------------------------------------------------------
# Name:        Editing TimeProfiles for e.g. Bus-Lanes
# Purpose:
#
# Author:      mape
#
# Created:     31/05/2021
# Copyright:   (c) mape 2021
# Licence:     <your licence>
#-------------------------------------------------------------------------------

import win32com.client.dynamic
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import time
import numpy as np
import pandas as pd
start_time = time.time()

from pathlib import Path
f = open(Path.home() / 'python32' / 'python_dir.txt', mode='r')
for i in f: path = i
path = Path.joinpath(Path(path),'VISUM_Tools','TimeProfiles.txt')
f = path.read_text().split('\n')

#--Path--#
Network = f[0]
Excel = f[1]

def LaneTest(LineRoute,StartHP,EndHP):
    i = 0
    index_start = 500
    for LRItem in LineRoute.LineRouteItems:
        if LRItem.AttValue("StopPointNo") == StartHP and LRItem.AttValue("IsRoutePoint") == True: 
            i,e = 1,0.0
            index_start = LRItem.AttValue("Index")
        if LRItem.AttValue(r'OutLink\Busspur') == None: return e
        if LRItem.AttValue("StopPointNo") == EndHP and LRItem.AttValue("IsRoutePoint") == True:
            if LRItem.AttValue("Index") > index_start: return e
        if i == 1 and LRItem.AttValue(r'OutLink\Busspur') > 0: e+=LRItem.AttValue(r'OutLink\Length')
        
def VISUM_open(Net):
    VISUM = win32com.client.dynamic.Dispatch("Visum.Visum.20")
    VISUM.loadversion(Net)
    VISUM.Filters.InitAll()
    return VISUM       

def sumTT(sheet):
    TT = pd.DataFrame(sheet.values)
    TT = TT.iloc[1:]
    TT[16] = TT.groupby([1,2,3,4])[13].transform('sum')
    TT[15] = TT[14] - TT[16]
    i = 2
    for row in TT.itertuples():
        sheet.cell(i,15,round((sheet.cell(i,15).value/60),1))
        sheet.cell(i,16,round((row[16]/60),1))
        sheet.cell(i,17,round((row[17]/60),1))
        i+=1

def exlwriter(TPItem,Share,Index,minTT):
    row = sheet.max_row + 1
    col_range = sheet.max_column
    newTT = minTT
    
    sheet.cell(row,1,row)
    sheet.cell(row,2,TPItem.AttValue("LineName"))
    sheet.cell(row,3,TPItem.AttValue("LineRouteName"))
    sheet.cell(row,4,(TPItem.AttValue("DirectionCode")+":"+
                      str(TPItem.AttValue("TimeProfileID"))+":"+
                      str(TPItem.AttValue("Index"))))
    sheet.cell(row,5,TPItem.AttValue("TimeProfileName"))
    sheet.cell(row,6,Index)
    sheet.cell(row,7,TPItem.AttValue(r'Max:UsedLineRouteItems\OutLink\Busspur'))
    sheet.cell(row,8,TPItem.AttValue(r'LineRouteItem\StopPoint\Name'))
    sheet.cell(row,9,TPItem.AttValue(r'NextTimeProfileItem\LineRouteItem\StopPoint\Name'))
    sheet.cell(row,10,Share)
    sheet.cell(row,11,TPItem.AttValue(r'PostRunTime'))
    if Share == 1:
        Bonus = 0.85
        if newTT > TPItem.AttValue(r'PostRunTime'): newTT = TPItem.AttValue(r'PostRunTime')
        sheet.cell(row,12,newTT)
        sheet.cell(row,13,(newTT*Bonus))
        sheet.cell(row,14,TPItem.AttValue(r'PostRunTime')-(newTT*Bonus))
    else:
        Bonus = 1-(0.15*Share)
        sheet.cell(row,12,TPItem.AttValue(r'PostRunTime'))
        sheet.cell(row,13,(TPItem.AttValue(r'PostRunTime')*Bonus))
        sheet.cell(row,14,TPItem.AttValue(r'PostRunTime')-(TPItem.AttValue(r'PostRunTime')*Bonus))
    sheet.cell(row,15,TPItem.AttValue(r'TimeProfile\Sum:TimeProfileItems\PostRunTime'))
    
    #color
    if Share == 1:
        greenFill = PatternFill(start_color='a8ff65',end_color='a8ff65',fill_type='solid')
        for col in range(1,col_range+1):sheet.cell(row,col).fill = greenFill
    if Share < 1 and Share >0.8:
        orangeFill = PatternFill(start_color='ffc965',end_color='ffc965',fill_type='solid')
        for col in range(1,col_range+1):sheet.cell(row,col).fill = orangeFill
    if minTT == 0: sheet.cell(row,12).fill = PatternFill(start_color='FFFF0000',end_color='FFFF0000',fill_type='solid')
  
def TProfiles(VISUM):
    FilterTP = VISUM.Filters.LineGroupFilter()
    FilterTP.UseFilterForVehJourneyItems = True
    FilterTP = FilterTP.VehJourneyItemFilter()
    FilterTP.AddCondition("OP_NONE",False,"Dep","GreaterEqualVal",32400) #32400 seconds == 9h
    FilterTP.AddCondition("OP_AND",False,"Dep","LessEqualVal",50400) #50400 seconds == 14h
    TProfiles = pd.DataFrame(np.array(VISUM.Net.TimeProfileItems.GetMultipleAttributes([r'LineRouteItem\StopPointNo',
                        r'NextTimeProfileItem\LineRouteItem\StopPointNo',
                        r'PostRunTime', r'MinActive:VehJourneyItems\Dep'])))  
    return TProfiles
  

#--Excel--#
wb = Workbook()
sheet = wb.active
sheet.cell(1,1,"ID")
sheet.cell(1,2,"Linie")
sheet.cell(1,3,"Route")
sheet.cell(1,4,"Richtung")
sheet.cell(1,5,"FZprofil")
sheet.cell(1,6,"Index")
sheet.cell(1,7,"Busspur")
sheet.cell(1,8,"vonHalt")
sheet.cell(1,9,"nachHalt")
sheet.cell(1,10,"BS_Anteil")
sheet.cell(1,11,"Fahrtzeit_sek")
sheet.cell(1,12,"min:Fahrtzeit_sek")
sheet.cell(1,13,"neu:Fahrtzeit_sek")
sheet.cell(1,14,"diff:Fahrtzeit_sek")
sheet.cell(1,15,"FZGesamt_min")
sheet.cell(1,16,"neu:FZGesamt_min")
sheet.cell(1,17,"diff:FZGesamt_min")

#VISUM
VISUM = VISUM_open(Network)
TProfiles = TProfiles(VISUM)

#--Line loop--#
for Line in VISUM.Net.Lines:
    # Line = VISUM.Net.Lines.ItemByKey("3")   
    if Line.AttValue("TSysCode") != "Bus": continue
    print(Line.AttValue("Name"))
    for LineRoute in Line.LineRoutes:
        for TimeProfile in LineRoute.TimeProfiles:
            Index = 0
            for TPItem in TimeProfile.TimeProfileItems:
                Index+=1
                
                StartHP = TPItem.AttValue(r'LineRouteItem\StopPointNo')
                EndHP = TPItem.AttValue(r'NextTimeProfileItem\LineRouteItem\StopPointNo')
                if TPItem.AttValue(r'Sum:UsedLineRouteItems\OutLink\Busspur') in [0, None]:continue
                
                LaneLength = LaneTest(LineRoute,StartHP,EndHP)
                Share = round(LaneLength/TPItem.AttValue(r'PostLength'),2)
                if Share < 0.25: continue
                minTT = TProfiles.loc[(TProfiles[0] == StartHP) & (TProfiles[1] == EndHP) & (TProfiles[3].notnull())].min()[2]
    
                #writer_to_excel
                exlwriter(TPItem, Share, Index, minTT)

#newTraveltime
sumTT(sheet)
  
#END
sheet.auto_filter.ref = sheet.dimensions
wb.save(Excel)
wb.close()
seconds = int(time.time() - start_time)
print("--finished after ",seconds,"seconds--") 